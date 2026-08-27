"""
미트웰 발주서 변환 - 핵심 처리 로직
스마트스토어 XLSX + 카페24 CSV → 미트웰 발주 XLSX
"""
import re
import io
import base64
import hashlib
import struct
import xml.etree.ElementTree as ET

import pandas as pd
import olefile
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend


# ──────────────────────────────────────────
# 고정 세트 구성표
# ──────────────────────────────────────────
SET_10_200G = [
    ("부채살(호주산)", "스테이크", "200g"),
    ("부채살(호주산)", "큐브",     "200g"),
    ("우삼겹(미국산)", "2mm",      "200g"),
    ("홍두깨살(호주산)", "스테이크", "200g"),
    ("홍두깨살(호주산)", "큐브",    "200g"),
    ("홍두깨살(호주산)", "슬라이스","200g"),
    ("우둔살(호주산)", "스테이크",  "200g"),
    ("우둔살(호주산)", "큐브",      "200g"),
    ("우둔살(호주산)", "슬라이스",  "200g"),
    ("지방제한(호주산)", "소고기 다짐육", "200g"),
]

SET_9_100G = [
    ("부채살(호주산)", "스테이크", "100g"),
    ("부채살(호주산)", "큐브",    "100g"),
    ("부채살(호주산)", "슬라이스","100g"),
    ("우둔살(호주산)", "슬라이스","100g"),
    ("우둔살(호주산)", "큐브",    "100g"),
    ("우둔살(호주산)", "스테이크","100g"),
    ("홍두깨살(호주산)", "슬라이스","100g"),
    ("홍두깨살(호주산)", "큐브",   "100g"),
    ("홍두깨살(호주산)", "스테이크","100g"),
]

SET_6_200G = [
    ("부채살(호주산)", "스테이크",       "200g"),
    ("부채살(호주산)", "큐브",           "200g"),
    ("우삼겹(미국산)", "2mm",            "200g"),
    ("홍두깨살(호주산)", "슬라이스",      "200g"),
    ("우둔살(호주산)", "슬라이스",        "200g"),
    ("지방제한(호주산)", "소고기 다짐육", "200g"),
]


# ──────────────────────────────────────────
# 유틸 함수
# ──────────────────────────────────────────
def normalize_cutting(cutting: str) -> str:
    cutting = cutting.strip()
    if "큐브" in cutting:
        return "큐브"
    if "통스테이크" in cutting or ("스테이크" in cutting and "큐브" not in cutting):
        return "스테이크"
    if "슬라이스" in cutting:
        return "슬라이스"
    if "2mm" in cutting:
        return "2mm"
    if "다짐육" in cutting:
        return "소고기 다짐육"
    return cutting


def normalize_buwi(buwi: str) -> str:
    """부위명 → 산지 포함 정규화"""
    buwi = buwi.strip()
    if "호주산" in buwi or "미국산" in buwi:
        return buwi
    if "우삼겹" in buwi:
        return "우삼겹(미국산)"
    if "지방제한" in buwi or "다짐육" in buwi:
        return "지방제한(호주산)"
    return f"{buwi}(호주산)"


def build_product_name(buwi: str, cutting: str, weight: str) -> str:
    if "다짐육" in cutting:
        return f"{normalize_buwi(buwi)} {cutting} {weight}"
    return f"{normalize_buwi(buwi)} {normalize_cutting(cutting)} {weight}"


def parse_size_packs(text: str):
    """
    '100g X 10팩 소포장' 또는 '200g x 5팩' 같은 문자열에서
    (무게, 팩수) 추출
    """
    m = re.search(r'(\d+g)\s*[xX×]\s*(\d+)팩', text)
    if m:
        return m.group(1), int(m.group(2))
    return None, 1


def get_set_by_name(product_name: str, quantity: int):
    """상품명에서 세트 종류 판별 → [(품목명, 팩수), ...]"""
    name = str(product_name)
    # 9종을 먼저 체크 (맛보기 패키지가 9종·10종 둘 다 포함되어 있어 순서 중요)
    if "9종" in name:
        template = SET_9_100G
    elif "인기 6종" in name or "6종" in name:
        template = SET_6_200G
    elif "10종" in name or "맛보기 패키지" in name:
        template = SET_10_200G
    else:
        return None  # 인식 불가

    result = []
    for b, c, w in template:
        result.append((build_product_name(b, c, w), quantity))
    return result


# ──────────────────────────────────────────
# 스마트스토어 파일 복호화
# ──────────────────────────────────────────
def decrypt_smartstore(path: str, password: str = "1111") -> io.BytesIO:
    """ECMA-376 Agile 방식으로 암호화된 스마트스토어 XLSX 복호화"""
    with open(path, "rb") as f:
        ole = olefile.OleFileIO(f)
        enc_info_raw = ole.openstream("EncryptionInfo").read()
        enc_pkg = ole.openstream("EncryptedPackage").read()

    root = ET.fromstring(enc_info_raw[8:])
    ns_e = "http://schemas.microsoft.com/office/2006/encryption"
    ns_p = "http://schemas.microsoft.com/office/2006/keyEncryptor/password"
    kd = root.find(f"{{{ns_e}}}keyData")
    ke = root.find(f".//{{{ns_p}}}encryptedKey")

    data_salt   = base64.b64decode(kd.get("saltValue"))
    key_salt    = base64.b64decode(ke.get("saltValue"))
    enc_key_val = base64.b64decode(ke.get("encryptedKeyValue"))
    spin_count  = int(ke.get("spinCount"))
    key_len     = int(ke.get("keyBits")) // 8
    block_size  = int(ke.get("blockSize"))

    def aes_cbc(key, data, iv):
        c = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
        d = c.decryptor()
        return d.update(data) + d.finalize()

    h = hashlib.sha512(key_salt + password.encode("utf-16-le")).digest()
    for i in range(spin_count):
        h = hashlib.sha512(struct.pack("<I", i) + h).digest()

    block_key = bytes([0x14, 0x6e, 0x0b, 0xe7, 0xab, 0xac, 0xd0, 0xd6])
    derived = hashlib.sha512(h + block_key).digest()[:key_len]
    secret_key = aes_cbc(derived, enc_key_val, key_salt)[:key_len]

    total = struct.unpack("<Q", enc_pkg[:8])[0]
    enc = enc_pkg[8:]
    seg_size = 4096
    out = b""

    for i in range((len(enc) + seg_size - 1) // seg_size):
        chunk = enc[i * seg_size : (i + 1) * seg_size]
        iv = hashlib.sha512(data_salt + struct.pack("<I", i)).digest()[:block_size]
        pad = (block_size - len(chunk) % block_size) % block_size
        chunk += b"\x00" * pad
        out += aes_cbc(secret_key, chunk, iv)

    return io.BytesIO(out[:total])


# ──────────────────────────────────────────
# 카페24 CSV 처리
# ──────────────────────────────────────────
def parse_cafe24_row(row: pd.Series) -> list:
    """
    카페24 주문 1행 → [(품목명, 팩수, 사은품여부)] 리스트
    사은품여부: True = 이 행은 사은품 카운트에 포함
    """
    product_name = str(row.get("주문상품명", ""))
    option       = str(row.get("옵션", ""))
    quantity     = int(row.get("수량", 1) or 1)

    # 옵션 없으면 고정 세트
    if option == "nan" or not option.strip():
        items = get_set_by_name(product_name, quantity)
        if items:
            return [(p, n, True) for p, n in items]
        return []

    # 인기 6종 종합세트 (골라담기 특수 옵션)
    if "인기 6종 종합세트" in option:
        result = []
        for b, c, w in SET_6_200G:
            result.append((build_product_name(b, c, w), quantity, True))
        return result

    # 옵션 파싱 (세미콜론 구분) - 키 공백 정규화
    parts = {}
    for part in option.split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            # 공백 정규화: '컷팅 방식' → '컷팅방식'
            key = k.strip().replace(" ", "")
            parts[key] = v.strip()

    buwi    = parts.get("소고기부위종류", parts.get("소고기종류", ""))
    cutting = parts.get("컷팅방식", "")
    sopojang = parts.get("소포장선택", "")

    if sopojang and sopojang != "nan":
        # 골라담기: 소포장 선택에 무게·팩수
        weight, packs = parse_size_packs(sopojang)
        product = build_product_name(buwi, cutting, weight or "")
        return [(product, packs * quantity, True)]
    else:
        # 단품 (X N팩): 컷팅방식 필드에 무게·팩수 포함
        weight, packs = parse_size_packs(cutting)
        if weight:
            cut_clean = re.sub(r'\d+g.*', '', cutting).strip()
            product = build_product_name(buwi, cut_clean, weight)
            return [(product, packs * quantity, True)]
        return []


def process_cafe24(csv_path: str) -> pd.DataFrame:
    """카페24 CSV → 발주 행 DataFrame"""
    for enc in ["utf-8-sig", "euc-kr", "cp949"]:
        try:
            df = pd.read_csv(csv_path, encoding=enc)
            break
        except Exception:
            continue

    rows_out = []
    prev_key = None
    saeum_count = {}

    for _, row in df.iterrows():
        items = parse_cafe24_row(row)
        if not items:
            continue

        name    = str(row.get("수령인", ""))
        addr    = str(row.get("주소", ""))
        phone1  = str(row.get("전화번호", "") or row.get("핸드폰", ""))
        phone2  = str(row.get("핸드폰", "") or phone1)
        msg     = str(row.get("비고", "")) if str(row.get("비고", "")) != "nan" else ""

        # 전화번호 정리
        phone1 = phone1 if phone1 != "nan" else phone2
        phone2 = phone2 if phone2 != "nan" else phone1

        cust_key = (name, addr)
        saeum_count[cust_key] = saeum_count.get(cust_key, 0) + sum(1 for _, _, g in items if g)

        for product, packs, _ in items:
            rows_out.append({
                "_name": name,
                "_addr": addr,
                "_phone1": phone1,
                "_phone2": phone2,
                "_msg": msg,
                "_product": product,
                "_packs": packs,
                "_source": "cafe24",
                "_cust_key": cust_key,
            })

    return pd.DataFrame(rows_out), saeum_count


# ──────────────────────────────────────────
# 스마트스토어 XLSX 처리
# ──────────────────────────────────────────
SKIP_PRODUCTS = ["시즈닝", "시즌닝", "와사비", "허브솔트"]  # 사은품성 제품


def is_skip_product(product_name: str) -> bool:
    return any(k in product_name for k in SKIP_PRODUCTS)


def parse_ss_option(option: str, quantity: int):
    """
    스마트스토어 옵션정보 파싱
    '소고기 부위 종류: 부채살(호주산) / 컷팅방식: 슬라이스 / 소포장 선택: 100g X 10팩 소포장'
    '소고기 1팩 추가: 부채살 큐브스테이크 200g'
    """
    if not option or str(option) == "nan":
        return None

    option = str(option)

    # 슬래시(/) 구분 (골라담기형)
    if "/" in option:
        parts = {}
        for part in option.split("/"):
            if ":" in part:
                k, v = part.split(":", 1)
                parts[k.strip()] = v.strip()

        buwi    = parts.get("소고기 부위 종류", "")
        cutting = parts.get("컷팅방식", "")
        sopojang= parts.get("소포장 선택", "")

        # 인기 6종 종합세트 → 개별 행으로 펼침
        if "인기 6종" in cutting or "6종 종합세트" in cutting:
            return [(build_product_name(b, c, w), quantity, True) for b, c, w in SET_6_200G]

        weight, packs = parse_size_packs(sopojang)

        # 소포장 선택 없이 컷팅방식에 무게·팩수가 포함된 경우
        # 예: '컷팅방식: 슬라이스 200g X 5팩'
        if not weight:
            weight, packs = parse_size_packs(cutting)
            if weight:
                cutting = re.sub(r'\s*\d+g\s*[xX×]\s*\d+팩.*', '', cutting, flags=re.IGNORECASE).strip()

        product = build_product_name(buwi, cutting, weight or "")
        return [(product, packs * quantity, True)]

    # 돼지고기 추가 옵션 (목살 등)
    # 예: '돼지고기도 함께(배송비 절약): 구이용 목살 150g X 6팩 = 900g'
    if "목살" in option:
        detail = option.split(":", 1)[-1].strip() if ":" in option else option
        m_w = re.search(r'(\d+g)', detail)
        m_p = re.search(r'[xX×]\s*(\d+)팩', detail)
        weight = m_w.group(1) if m_w else ""
        packs  = int(m_p.group(1)) if m_p else quantity
        return [(f"목살 {weight} 서비스", packs, True)]

    # 단품 추가형 (1팩 추가: ...)
    m = re.search(r'추가:\s*(.+)', option)
    if m:
        detail = m.group(1).strip()
        # detail 예: '부채살 큐브스테이크 200g'
        wm = re.search(r'(\d+g)', detail)
        weight = wm.group(1) if wm else ""
        name_part = detail[:wm.start()].strip() if wm else detail

        # 부위와 컷팅 분리 (마지막 단어가 컷팅, 나머지가 부위)
        tokens = name_part.split()
        if len(tokens) >= 2:
            buwi = " ".join(tokens[:-1])
            cutting = tokens[-1]
        else:
            buwi = name_part
            cutting = ""

        product = build_product_name(buwi, cutting, weight)
        return [(product, quantity, True)]

    return None


def process_smartstore(xlsx_path: str, password: str = "1111") -> tuple:
    """스마트스토어 XLSX → 발주 행 DataFrame"""
    buf = decrypt_smartstore(xlsx_path, password)

    from openpyxl import load_workbook
    wb = load_workbook(buf)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Row0 = 안내문, Row1 = 헤더, Row2+ = 데이터
    headers = all_rows[1]
    h = {v: i for i, v in enumerate(headers) if v}

    rows_out = []
    saeum_count = {}

    for row in all_rows[2:]:
        def get(col):
            idx = h.get(col)
            return row[idx] if idx is not None else None

        product_name = str(get("상품명") or "")
        if is_skip_product(product_name):
            continue

        option   = get("옵션정보")
        quantity = int(get("수량") or 1)
        name     = str(get("수취인명") or "")
        addr     = str(get("통합배송지") or "")
        phone1   = str(get("수취인연락처1") or "")
        phone2   = str(get("수취인연락처2") or phone1)
        msg      = str(get("배송메세지") or "")
        if msg == "None":
            msg = ""

        cust_key = (name, addr)

        # 옵션 파싱 시도
        items = parse_ss_option(option, quantity)

        # 옵션 없으면 고정 세트 또는 알 수 없음
        if items is None:
            items = get_set_by_name(product_name, quantity)
            if items is None:
                # 파싱 불가 옵션 → 옵션 문자열을 품목명으로 그대로 포함
                opt_str = str(option) if option and str(option) != "nan" else ""
                fallback = opt_str if opt_str else product_name
                items = [(fallback, quantity, True)] if fallback else []
            else:
                items = [(p, n, True) for p, n in items]

        saeum_count[cust_key] = saeum_count.get(cust_key, 0) + sum(1 for _, _, g in items if g)

        for product, packs, _ in items:
            rows_out.append({
                "_name":    name,
                "_addr":    addr,
                "_phone1":  phone1,
                "_phone2":  phone2 if str(phone2) != "None" else phone1,
                "_msg":     msg,
                "_product": product,
                "_packs":   packs,
                "_source":  "smartstore",
                "_cust_key": cust_key,
            })

    return pd.DataFrame(rows_out), saeum_count


# ──────────────────────────────────────────
# 결과 Excel 생성
# ──────────────────────────────────────────
def build_output(ss_rows: pd.DataFrame, c24_rows: pd.DataFrame,
                 ss_saeum: dict, c24_saeum: dict) -> pd.DataFrame:
    """두 소스 합산 → 미트웰 발주서 DataFrame"""

    all_rows = pd.concat([ss_rows, c24_rows], ignore_index=True)
    saeum_map = {}
    for k, v in ss_saeum.items():
        saeum_map[k] = saeum_map.get(k, 0) + v
    for k, v in c24_saeum.items():
        saeum_map[k] = saeum_map.get(k, 0) + v

    output_rows = []
    first_per_customer = set()

    for _, r in all_rows.iterrows():
        cust_key = r["_cust_key"]
        is_first = cust_key not in first_per_customer
        if is_first:
            first_per_customer.add(cust_key)

        saeum_cnt = saeum_map.get(cust_key, 1) if is_first else None

        output_rows.append({
            "수하인명":        r["_name"],
            "주소":           r["_addr"],
            "수하인전화번호":  r["_phone1"],
            "수하인핸드폰번호":r["_phone2"],
            "박스수량":        1,
            "택배운임":        3000,
            "운임구분":        "선불",
            "품목명":          r["_product"],
            "팩 개수":         r["_packs"],
            "배송메세지":      r["_msg"] if is_first else None,
            "사은품":          "시즌닝 3종세트" if is_first else None,
            "사은품갯수":      saeum_cnt,
            "**비고란":        None,
            "택배사":          None,
            "송장번호":        None,
        })

    return pd.DataFrame(output_rows)


def save_output_excel(df: pd.DataFrame, out_path: str):
    """결과 DataFrame → XLSX 저장"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "발주서"

    cols = list(df.columns)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 데이터
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = None
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 컬럼 너비 자동
    col_widths = [12, 40, 16, 16, 8, 10, 8, 30, 8, 25, 14, 8, 10, 10, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    wb.save(out_path)
